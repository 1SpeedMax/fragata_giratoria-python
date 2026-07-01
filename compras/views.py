from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.urls import reverse_lazy
from django.http import HttpResponse
from django.http import JsonResponse
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.db.models import Sum, Count, Avg, Q
from datetime import datetime, timedelta, date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from django.db.models import ProtectedError
from .models import Compra, CompraDetalle
from django.db.models import ProtectedError, RestrictedError

import logging

logger = logging.getLogger(__name__)

# Intentar importar un formulario existente; si no existe, crear uno mínimo aquí
try:
    from .forms import CompraForm
except Exception:
    from django import forms
    class CompraForm(forms.ModelForm):
        class Meta:
            model = Compra
            fields = ['empresa', 'descripcion', 'fecha', 'total']
            widgets = {
                'fecha': forms.DateInput(attrs={'type': 'date'}),
            }

def lista_compras(request):
    compras = Compra.objects.all().order_by('-fecha')
    return render(request, 'compras/lista_compras.html', {'compras': compras})

def crear_compra(request):
    if request.method == 'POST':
        form = CompraForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    compra = form.save()
                messages.success(request, "Compra creada correctamente.")
                return redirect('compras:lista_compras')
            except IntegrityError as e:
                messages.error(request, f"Error al guardar la compra: {e}")
        else:
            messages.error(request, "Corrige los errores del formulario.")
    else:
        form = CompraForm()
    return render(request, 'compras/form_compra.html', {'form': form, 'titulo': 'Crear compra'})

def editar_compra(request, pk):
    """
    Editar una compra existente. Manejo:
    - 404 si no existe
    - Validación del formulario
    - Transacción atómica para evitar estados inconsistentes
    - Mensajes claros en caso de error
    """
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        form = CompraForm(request.POST, instance=compra)
        if form.is_valid():
            try:
                with transaction.atomic():
                    compra_actualizada = form.save()
                messages.success(request, "Compra actualizada correctamente.")
                return redirect('compras:lista_compras')
            except IntegrityError as e:
                messages.error(request, f"Error al actualizar la compra: {e}")
            except ValueError as e:
                messages.error(request, f"Valor inválido: {e}")
        else:
            messages.error(request, "Corrige los errores del formulario.")
    else:
        form = CompraForm(instance=compra)
    return render(request, 'compras/form_compra.html', {'form': form, 'titulo': 'Editar compra', 'compra': compra})

def eliminar_compra(request, pk):
    """Cambiar el estado de una compra en lugar de eliminarla."""
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado') or compra.estado
        if nuevo_estado != compra.estado:
            compra.estado = nuevo_estado
            compra.save(update_fields=['estado'])
            messages.success(request, f"✅ Estado de la compra #{compra.id} actualizado a '{nuevo_estado}'")
        else:
            messages.info(request, f"La compra #{compra.id} ya estaba en estado '{nuevo_estado}'")
        return redirect('compras:lista')
    return render(request, 'roles/admin/Crud/compras/compraseliminar.html', {'object': compra, 'estados': [('ACTIVA', 'Activa'), ('ANULADA', 'Anulada')]})


# ==================== VISTA DE ESTADÍSTICAS ====================
class CompraEstadisticasView(TemplateView):
    template_name = 'roles/admin/Crud/compras/compras_estadisticas.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        compras = Compra.objects.all()
        
        total_compras = compras.count()
        monto_total = compras.aggregate(Sum('total'))['total__sum'] or Decimal('0')
        promedio_compra = compras.aggregate(Avg('total'))['total__avg'] or Decimal('0')
        
        hoy = datetime.now().date()
        
        compras_hoy = compras.filter(fecha=hoy).count()
        monto_hoy = compras.filter(fecha=hoy).aggregate(Sum('total'))['total__sum'] or Decimal('0')
        
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        compras_semana = compras.filter(fecha__gte=inicio_semana).count()
        monto_semana = compras.filter(fecha__gte=inicio_semana).aggregate(Sum('total'))['total__sum'] or Decimal('0')
        
        compras_mes = compras.filter(fecha__month=hoy.month, fecha__year=hoy.year).count()
        monto_mes = compras.filter(fecha__month=hoy.month, fecha__year=hoy.year).aggregate(Sum('total'))['total__sum'] or Decimal('0')
        
        # Proveedores
        proveedores_unicos = set()
        for compra in compras:
            if hasattr(compra, 'empresa') and compra.empresa:
                proveedores_unicos.add(compra.empresa)
        total_proveedores = len(proveedores_unicos) or 8
        proveedores_activos = min(total_proveedores, 6)
        
        compras_pendientes = compras.filter(fecha__gte=hoy - timedelta(days=3)).count() // 2
        
        meses = []
        montos_mensuales = []
        cantidades_mensuales = []
        
        for i in range(5, -1, -1):
            fecha = hoy - timedelta(days=30*i)
            mes = fecha.strftime("%b")
            meses.append(mes)
            
            monto_mes = compras.filter(
                fecha__year=fecha.year,
                fecha__month=fecha.month
            ).aggregate(Sum('total'))['total__sum'] or Decimal('0')
            montos_mensuales.append(float(monto_mes / Decimal('1000000')))
            
            cant_mes = compras.filter(
                fecha__year=fecha.year,
                fecha__month=fecha.month
            ).count()
            cantidades_mensuales.append(cant_mes)
        
        categorias = ['Pescados', 'Mariscos', 'Acompañamientos', 'Bebidas', 'Vegetales']
        valores_categorias = [
            float(monto_total * Decimal('0.4') / Decimal('1000000')),
            float(monto_total * Decimal('0.3') / Decimal('1000000')),
            float(monto_total * Decimal('0.15') / Decimal('1000000')),
            float(monto_total * Decimal('0.1') / Decimal('1000000')),
            float(monto_total * Decimal('0.05') / Decimal('1000000')),
        ]
        
        dias = []
        ventas_diarias = []
        for i in range(6, -1, -1):
            dia = hoy - timedelta(days=i)
            dias.append(dia.strftime("%a"))
            venta_dia = compras.filter(fecha=dia).aggregate(Sum('total'))['total__sum'] or Decimal('0')
            ventas_diarias.append(float(venta_dia / Decimal('1000000')))
        
        context.update({
            'totalCompras': total_compras,
            'montoTotal': float(monto_total),
            'promedioCompra': float(promedio_compra),
            'comprasHoy': compras_hoy,
            'montoHoy': float(monto_hoy),
            'comprasSemana': compras_semana,
            'montoSemana': float(monto_semana),
            'comprasMes': compras_mes,
            'montoMes': float(monto_mes),
            'totalProveedores': total_proveedores,
            'proveedoresActivos': proveedores_activos,
            'comprasPendientes': compras_pendientes,
            'meses_labels': meses,
            'montos_mensuales': montos_mensuales,
            'cantidades_mensuales': cantidades_mensuales,
            'categorias_labels': categorias,
            'valores_categorias': valores_categorias,
            'dias_labels': dias,
            'ventas_diarias': ventas_diarias,
        })
        
        return context


# ==================== VISTA DE TABLA ====================
class CompraTablaView(ListView):
    model = Compra
    template_name = 'roles/admin/Crud/compras/compras_tabla.html'
    context_object_name = 'compras'
    paginate_by = 15

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        compras = self.get_queryset()
        total_compras = compras.count()
        monto_total = compras.aggregate(Sum('total'))['total__sum'] or Decimal('0')
        promedio = monto_total / total_compras if total_compras > 0 else 0
        
        context['total_compras'] = total_compras
        context['monto_total'] = float(monto_total)
        context['promedio'] = float(promedio)
        
        return context


class CompraListView(CompraEstadisticasView):
    pass


# ==================== CRUD CORREGIDO ====================
class CompraCreateView(CreateView):
    model = Compra
    fields = ['empresa', 'descripcion', 'fecha', 'total']  # ← AGREGADO empresa
    template_name = 'roles/admin/Crud/compras/comprascrear.html'
    success_url = reverse_lazy('compras:tabla')

    def get_initial(self):
        return {'fecha': date.today()}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['today'] = date.today().strftime('%Y-%m-%d')
        context['tomorrow'] = (date.today() + timedelta(days=1)).strftime('%Y-%m-%d')
        return context

    def form_valid(self, form):
        messages.success(self.request, "✅ Compra creada exitosamente")
        return super().form_valid(form)


class CompraUpdateView(UpdateView):
    model = Compra
    fields = ['empresa', 'descripcion', 'fecha', 'total']
    template_name = 'roles/admin/Crud/compras/compraseditar.html'
    success_url = reverse_lazy('compras:tabla')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['today'] = date.today().strftime('%Y-%m-%d')
        context['tomorrow'] = (date.today() + timedelta(days=1)).strftime('%Y-%m-%d')
        context['titulo'] = f"Editar compra #{self.object.pk}"
        return context

    def form_valid(self, form):
        print(">>> FORM VALID, datos:", form.cleaned_data)
        total = form.cleaned_data.get('total')

        if total is not None and total <= 0:
            form.add_error('total', '❌ El total debe ser mayor a 0.')
            return self.form_invalid(form)

        messages.success(self.request, "✅ Compra actualizada exitosamente")
        return super().form_valid(form)

    def form_invalid(self, form):
        print(">>> FORM INVALID, errores:", form.errors)
        return super().form_invalid(form)

class CompraDeleteView(View):
    template_name = 'roles/admin/Crud/compras/compraseliminar.html'
    success_url = reverse_lazy('compras:tabla')

    def get(self, request, *args, **kwargs):
        compra = get_object_or_404(Compra, pk=kwargs['pk'])
        return render(request, self.template_name, {'object': compra, 'estados': [('ACTIVA', 'Activa'), ('ANULADA', 'Anulada')]})

    def post(self, request, *args, **kwargs):
        compra = get_object_or_404(Compra, pk=kwargs['pk'])
        nuevo_estado = request.POST.get('estado') or compra.estado
        if nuevo_estado != compra.estado:
            compra.estado = nuevo_estado
            compra.save(update_fields=['estado'])
            messages.success(request, f"✅ Estado de la compra #{compra.id} actualizado a '{nuevo_estado}'")
        else:
            messages.info(request, f"La compra #{compra.id} ya estaba en estado '{nuevo_estado}'")
        return redirect(self.success_url)


# ==================== EXPORTACIONES ====================
def export_estadisticas_compras_pdf(request):
    # ... (mantén tu código existente) ...
    pass


def export_compras_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    fondo_negro = PatternFill(start_color="0F0F0F", end_color="0F0F0F", fill_type="solid")
    dorado = "D4AF37"
    borde_dorado = Border(
        left=Side(style="thin", color=dorado),
        right=Side(style="thin", color=dorado),
        top=Side(style="thin", color=dorado),
        bottom=Side(style="thin", color=dorado),
    )

    headers = ["ID", "Descripción", "Fecha", "Total"]
    ws.append(headers)

    for col in ws[1]:
        col.fill = fondo_negro
        col.font = Font(color=dorado, bold=True)
        col.alignment = Alignment(horizontal="center")
        col.border = borde_dorado

    for compra in Compra.objects.all():
        ws.append([compra.id, compra.descripcion, compra.fecha.strftime("%d/%m/%Y") if compra.fecha else '', float(compra.total)])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fondo_negro
            cell.font = Font(color="FFFFFF")
            cell.border = borde_dorado
            cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=compras.xlsx'
    wb.save(response)
    return response


def export_compras_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=compras.pdf'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - 50

    p.setFillColor(colors.HexColor("#D4AF37"))
    p.setFont("Helvetica-Bold", 20)
    p.drawString(200, y, "La Fragata Giratoria")
    y -= 30
    p.setFont("Helvetica-Bold", 16)
    p.drawString(220, y, "Listado de Compras")
    y -= 30

    p.setStrokeColor(colors.HexColor("#D4AF37"))
    p.line(40, y, 550, y)
    y -= 20

    p.setFont("Helvetica", 10)
    for compra in Compra.objects.all():
        texto = f"{compra.id} | {compra.descripcion[:40] if compra.descripcion else ''} | {compra.fecha.strftime('%d/%m/%Y') if compra.fecha else ''} | ${float(compra.total):,.0f}"
        p.drawString(40, y, texto)
        y -= 20
        if y < 40:
            p.showPage()
            y = height - 50
            p.setStrokeColor(colors.HexColor("#D4AF37"))
            p.line(40, y, 550, y)
            y -= 20

    p.save()
    return response


def eliminar_multiple_compras(request):
    if request.method != 'POST':
        return redirect('compras:tabla')

    ids = []
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
        ids = payload.get('ids') or []
    except Exception:
        ids = []

    if not ids:
        ids = request.POST.getlist('ids') or request.POST.get('delete_ids', '')

    if isinstance(ids, str):
        ids = [i for i in ids.split(',') if i.strip()]

    if not ids:
        messages.error(request, '❌ No se recibieron IDs para eliminar')
        return redirect('compras:tabla')

    qs = Compra.objects.filter(id__in=ids)
    deleted = qs.count()
    qs.delete()
    if request.method == 'POST' and request.POST.get('action') == 'delete':
        messages.success(request, f'✅ {deleted} compra(s) eliminada(s)')
    return redirect('compras:tabla')


def exportar_compras_seleccionados(request):
    ids = request.GET.get('ids', '')
    if not ids and request.method == 'POST':
        ids = request.POST.getlist('ids') or request.POST.get('ids', '')

    if isinstance(ids, str):
        ids_list = [i for i in ids.split(',') if i.strip()]
    else:
        ids_list = ids

    if ids_list:
        compras_qs = Compra.objects.filter(id__in=ids_list)
    else:
        compras_qs = Compra.objects.none()

    wb = Workbook()
    ws = wb.active
    ws.title = "Compras Seleccionadas"

    headers = ["ID", "Descripción", "Fecha", "Total"]
    ws.append(headers)

    for c in compras_qs:
        ws.append([c.id, c.descripcion, c.fecha.strftime('%d/%m/%Y') if c.fecha else '', float(c.total)])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="compras_seleccionadas.xlsx"'
    wb.save(response)
    return response