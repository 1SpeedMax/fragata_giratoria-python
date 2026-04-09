from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q, Count
from django.core.paginator import Paginator 
from datetime import datetime
from .models import MetodoPago
from django.views.generic import ListView

# ==================== LISTA DE MÉTODOS CON PAGINACIÓN ====================
def lista_metodos(request):
    """Vista para listar todos los métodos de pago con paginación"""
    
    # Obtener todos los métodos
    metodos_list = MetodoPago.objects.all().order_by('id_metodo_pago')
    
    # Búsqueda (opcional)
    search_query = request.GET.get('search', '')
    if search_query:
        metodos_list = metodos_list.filter(
            Q(nombre_metodo__icontains=search_query) |
            Q(descripcion__icontains=search_query)
        )
    
    # ===== PAGINACIÓN =====
    paginator = Paginator(metodos_list, 5)  # 5 items por página para probar
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_metodos = MetodoPago.objects.count()
    con_descripcion = MetodoPago.objects.exclude(descripcion__isnull=True).exclude(descripcion='').count()
    sin_descripcion = total_metodos - con_descripcion
    
    context = {
        'metodos': page_obj,  # Los métodos paginados
        'total_metodos': total_metodos,
        'con_descripcion': con_descripcion,
        'sin_descripcion': sin_descripcion,
        # Variables para paginación
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'paginator': paginator,
    }
    return render(request, 'roles/admin/Crud/metodospago/metodos.html', context)


# ==================== ESTADÍSTICAS ====================
def estadisticas_metodos(request):
    """Vista de estadísticas de métodos de pago"""
    metodos = MetodoPago.objects.all()
    
    # Clasificación por tipo (basada en el nombre)
    efectivo = metodos.filter(nombre_metodo__icontains='efectivo').count()
    tarjeta = metodos.filter(nombre_metodo__icontains='tarjeta').count()
    digital = metodos.filter(
        Q(nombre_metodo__icontains='nequi') |
        Q(nombre_metodo__icontains='daviplata') |
        Q(nombre_metodo__icontains='transferencia')
    ).count()
    otros = metodos.count() - (efectivo + tarjeta + digital)
    
    # Contar por longitud de descripción
    con_descripcion = metodos.exclude(descripcion__isnull=True).exclude(descripcion='').count()
    sin_descripcion = metodos.count() - con_descripcion
    
    context = {
        'total_metodos': metodos.count(),
        'metodos_por_tipo': {
            'efectivo': efectivo,
            'tarjeta': tarjeta,
            'digital': digital,
            'otros': otros,
        },
        'con_descripcion': con_descripcion,
        'sin_descripcion': sin_descripcion,
        'ultima_actualizacion': datetime.now(),
    }
    return render(request, 'roles/admin/Crud/metodospago/estadisticas_metodospago.html', context)


# ==================== CREAR MÉTODO ====================
def crear_metodo(request):
    """Vista para crear un nuevo método de pago"""
    if request.method == 'POST':
        nombre = request.POST.get('nombreMetodo')
        descripcion = request.POST.get('descripcionMetodo')
        
        if nombre:
            if MetodoPago.objects.filter(nombre_metodo__iexact=nombre).exists():
                messages.error(request, f'❌ El método "{nombre}" ya existe')
            else:
                metodo = MetodoPago(
                    nombre_metodo=nombre,
                    descripcion=descripcion,
                )
                metodo.save()
                messages.success(request, f'✅ Método "{nombre}" creado exitosamente')
                return redirect('metodos_pago:lista')
        else:
            messages.error(request, '❌ El nombre es obligatorio')
    
    return render(request, 'roles/admin/Crud/metodospago/crear_metodo.html')

class MetodoPagoListView(ListView):
    model = MetodoPago
    template_name = 'ruta_de_tu_html.html'
    context_object_name = 'metodos'
    paginate_by = 10


# ==================== EDITAR MÉTODO ====================
def editar_metodo(request, pk):
    """Vista para editar un método de pago"""
    metodo = get_object_or_404(MetodoPago, id_metodo_pago=pk)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombreMetodo')
        descripcion = request.POST.get('descripcionMetodo')
        
        if nombre:
            if MetodoPago.objects.filter(nombre_metodo__iexact=nombre).exclude(id_metodo_pago=pk).exists():
                messages.error(request, f'❌ El método "{nombre}" ya existe')
            else:
                metodo.nombre_metodo = nombre
                metodo.descripcion = descripcion
                metodo.save()
                messages.success(request, f'✅ Método "{nombre}" actualizado')
                return redirect('metodos_pago:lista')
        else:
            messages.error(request, '❌ El nombre es obligatorio')
    
    context = {'metodo': metodo}
    return render(request, 'roles/admin/Crud/metodospago/editar_metodo.html', context)


# ==================== ELIMINAR MÉTODO ====================
def eliminar_metodo(request, pk):
    """Vista para eliminar un método de pago"""
    metodo = get_object_or_404(MetodoPago, id_metodo_pago=pk)
    
    if request.method == 'POST':
        nombre = metodo.nombre_metodo
        metodo.delete()
        messages.success(request, f'✅ Método "{nombre}" eliminado')
        return redirect('metodos_pago:lista')
    
    context = {'metodo': metodo}
    return render(request, 'roles/admin/Crud/metodospago/eliminar_metodo.html', context)


# ==================== EXPORTAR ESTADÍSTICAS A PDF ====================
def export_estadisticas_metodos_pdf(request):
    """Exportar estadísticas de métodos de pago a PDF con gráficos"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    metodos = MetodoPago.objects.all()
    
    # Clasificación por tipo
    efectivo = metodos.filter(nombre_metodo__icontains='efectivo').count()
    tarjeta = metodos.filter(nombre_metodo__icontains='tarjeta').count()
    digital = metodos.filter(
        Q(nombre_metodo__icontains='nequi') |
        Q(nombre_metodo__icontains='daviplata') |
        Q(nombre_metodo__icontains='transferencia')
    ).count()
    otros = metodos.count() - (efectivo + tarjeta + digital)
    
    con_descripcion = metodos.exclude(descripcion__isnull=True).exclude(descripcion='').count()
    sin_descripcion = metodos.count() - con_descripcion
    
    total_metodos = metodos.count()
    
    # Datos para gráficos
    tipos_labels = ['Efectivo', 'Tarjeta', 'Digital', 'Otros']
    tipos_data = [efectivo, tarjeta, digital, otros]
    
    descripciones_labels = ['Con Descripción', 'Sin Descripción']
    descripciones_data = [con_descripcion, sin_descripcion]
    
    # Porcentajes
    porcentaje_efectivo = (efectivo / total_metodos * 100) if total_metodos > 0 else 0
    porcentaje_tarjeta = (tarjeta / total_metodos * 100) if total_metodos > 0 else 0
    porcentaje_digital = (digital / total_metodos * 100) if total_metodos > 0 else 0
    porcentaje_otros = (otros / total_metodos * 100) if total_metodos > 0 else 0
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="estadisticas_metodos_pago.pdf"'
    
    # Crear documento PDF
    doc = SimpleDocTemplate(response, pagesize=landscape(letter),
                           leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#d4af37'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#bba163'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#f5d487'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    # Título
    story.append(Paragraph("LA FRAGATA GIRATORIA", title_style))
    story.append(Paragraph("Estadísticas de Métodos de Pago", subtitle_style))
    story.append(Spacer(1, 20))
    
    # Fecha de generación
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph(f"Fecha de generación: {fecha_actual}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # KPIs en tabla
    story.append(Paragraph("Métricas Principales", section_style))
    
    kpi_data = [
        ['Métrica', 'Valor', 'Porcentaje'],
        ['Total Métodos', str(total_metodos), '100%'],
        ['Métodos con Descripción', str(con_descripcion), f'{(con_descripcion / total_metodos * 100) if total_metodos > 0 else 0:.1f}%'],
        ['Métodos sin Descripción', str(sin_descripcion), f'{(sin_descripcion / total_metodos * 100) if total_metodos > 0 else 0:.1f}%'],
        ['Tipos de Métodos', str(len(tipos_labels)), '4 categorías'],
    ]
    
    kpi_table = Table(kpi_data, colWidths=[150, 120, 120])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4af37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 30))
    
    # Distribución por Tipo (gráfico de torta)
    story.append(Paragraph("Distribución por Tipo", section_style))
    
    if sum(tipos_data) > 0:
        drawing = Drawing(400, 220)
        pie = Pie()
        pie.x = 110
        pie.y = 30
        pie.width = 180
        pie.height = 180
        pie.data = tipos_data
        pie.labels = tipos_labels
        pie.slices.strokeWidth = 0.5
        colores_tipos = ['#10b981', '#3b82f6', '#f59e0b', '#8b5cf6']
        for i, color in enumerate(colores_tipos):
            if i < len(pie.slices):
                pie.slices[i].fillColor = colors.HexColor(color)
        drawing.add(pie)
        story.append(drawing)
    
    story.append(Spacer(1, 20))
    
    # Tabla de tipos
    tipos_table_data = [['Tipo', 'Cantidad', 'Porcentaje']]
    for label, value in zip(tipos_labels, tipos_data):
        porcentaje = (value / total_metodos * 100) if total_metodos > 0 else 0
        tipos_table_data.append([label, str(value), f'{porcentaje:.1f}%'])
    
    tipos_table = Table(tipos_table_data, colWidths=[150, 100, 100])
    tipos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4af37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
    ]))
    story.append(tipos_table)
    story.append(Spacer(1, 30))
    
    # Descripciones (gráfico de barras)
    story.append(Paragraph("Detalle de Descripciones", section_style))
    
    if max(descripciones_data) > 0:
        drawing2 = Drawing(500, 250)
        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 50
        bc.width = 400
        bc.height = 150
        bc.data = [descripciones_data]
        bc.categoryAxis.categoryNames = descripciones_labels
        bc.valueAxis.valueMin = 0
        bc.valueAxis.valueMax = max(descripciones_data) + (max(descripciones_data) * 0.2) if max(descripciones_data) > 0 else 10
        bc.bars[0].fillColor = colors.HexColor('#f5d487')
        bc.bars[0].strokeColor = colors.HexColor('#d4af37')
        drawing2.add(bc)
        story.append(drawing2)
    
    story.append(Spacer(1, 20))
    
    # Tabla de descripciones
    descripciones_table_data = [['Estado', 'Cantidad', 'Porcentaje']]
    for label, value in zip(descripciones_labels, descripciones_data):
        porcentaje = (value / total_metodos * 100) if total_metodos > 0 else 0
        descripciones_table_data.append([label, str(value), f'{porcentaje:.1f}%'])
    
    descripciones_table = Table(descripciones_table_data, colWidths=[150, 100, 100])
    descripciones_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4af37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
    ]))
    story.append(descripciones_table)
    story.append(Spacer(1, 30))
    
    # Información adicional
    story.append(Paragraph("Información Adicional", section_style))
    
    info_data = [
        ['Total Métodos de Pago', f'{total_metodos} métodos'],
        ['Tipos de Métodos', f'{len(tipos_labels)} categorías'],
        ['Método más común', max(zip(tipos_labels, tipos_data), key=lambda x: x[1])[0] if tipos_data else 'N/A'],
        ['Fecha de actualización', datetime.now().strftime("%d/%m/%Y")],
    ]
    
    info_table = Table(info_data, colWidths=[150, 250])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 30))
    
    # Pie de página
    story.append(Paragraph("Reporte generado automáticamente por el sistema La Fragata Giratoria", styles['Normal']))
    story.append(Paragraph("© 2025 - Todos los derechos reservados", styles['Normal']))
    
    doc.build(story)
    return response


# ==================== EXPORTAR A EXCEL ====================
def export_metodos_excel(request):
    """Exportar métodos de pago a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Métodos de Pago"
    
    fondo_negro = PatternFill(start_color="0F0F0F", end_color="0F0F0F", fill_type="solid")
    dorado = "D4AF37"
    borde_dorado = Border(
        left=Side(style="thin", color=dorado),
        right=Side(style="thin", color=dorado),
        top=Side(style="thin", color=dorado),
        bottom=Side(style="thin", color=dorado),
    )
    
    headers = ["ID", "Nombre", "Descripción"]
    ws.append(headers)
    
    for col in ws[1]:
        col.fill = fondo_negro
        col.font = Font(color=dorado, bold=True)
        col.alignment = Alignment(horizontal="center")
        col.border = borde_dorado
    
    for metodo in MetodoPago.objects.all():
        ws.append([
            metodo.id_metodo_pago,
            metodo.nombre_metodo,
            metodo.descripcion or ""
        ])
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fondo_negro
            cell.font = Font(color="FFFFFF")
            cell.border = borde_dorado
            cell.alignment = Alignment(horizontal="center")
    
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=metodos_pago.xlsx'
    wb.save(response)
    return response


# ==================== EXPORTAR A PDF ====================
def export_metodos_pdf(request):
    """Exportar métodos de pago a PDF"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=metodos_pago.pdf'
    
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - 50
    
    p.setFillColor(colors.HexColor("#D4AF37"))
    p.setFont("Helvetica-Bold", 20)
    p.drawString(200, y, "La Fragata Giratoria")
    y -= 30
    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, y, "Métodos de Pago")
    y -= 30
    
    p.setStrokeColor(colors.HexColor("#D4AF37"))
    p.line(40, y, 550, y)
    y -= 20
    
    p.setFont("Helvetica", 10)
    
    for metodo in MetodoPago.objects.all():
        texto = f"{metodo.id_metodo_pago} | {metodo.nombre_metodo} | {metodo.descripcion or ''}"
        p.drawString(40, y, texto[:80])
        y -= 15
        if y < 40:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica", 10)
    
    p.save()
    return response