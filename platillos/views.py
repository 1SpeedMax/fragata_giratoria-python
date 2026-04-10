from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.db.models import Q, Sum, Count, Avg, Min, Max, F
from django.utils import timezone
from datetime import datetime, timedelta
import calendar
from collections import Counter

from django.db.models import Prefetch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart

from .models import CategoriaPlatillo, Platillo
#Vista protegida
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render

#Definir admin para la protección de vitas
def es_admin(user):
    return user.is_staff

# ==================== MENU ====================
def menu(request):
    categorias = CategoriaPlatillo.objects.filter(activo=True).prefetch_related(
        Prefetch(
            'platillos',
            queryset=Platillo.objects.filter(disponible=True)
        )
    )

    return render(request, 'home/menu.html', {
        'categorias': categorias
    })

# ==================== CATEGORÍAS ====================

class CategoriaListView(ListView):
    model = CategoriaPlatillo
    template_name = 'roles/admin/Crud/platillos/categorias.html'
    context_object_name = 'categorias'
    paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_categorias'] = CategoriaPlatillo.objects.count()
        context['categorias_activas'] = CategoriaPlatillo.objects.filter(activo=True).count()
        return context


class CategoriaCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = CategoriaPlatillo
    fields = ['nombre', 'descripcion', 'emoji', 'activo', 'orden']
    template_name = 'roles/admin/Crud/platillos/categoria_crear.html'
    success_url = reverse_lazy('platillos:categorias')
    
    def test_func(self):
        return self.request.user.is_staff
    
    def form_valid(self, form):
        messages.success(self.request, f"✅ Categoría '{form.instance.nombre}' creada exitosamente")
        return super().form_valid(form)


class CategoriaUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CategoriaPlatillo
    fields = ['nombre', 'descripcion', 'emoji', 'activo', 'orden']
    template_name = 'roles/admin/Crud/platillos/categoria_editar.html'
    success_url = reverse_lazy('platillos:categorias')
    
    def test_func(self):
        return self.request.user.is_staff
    
    def form_valid(self, form):
        messages.success(self.request, f"✅ Categoría '{form.instance.nombre}' actualizada")
        return super().form_valid(form)


class CategoriaDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = CategoriaPlatillo
    template_name = 'roles/admin/Crud/platillos/categoria_eliminar.html'
    success_url = reverse_lazy('platillos:categorias')
    
    def test_func(self):
        return self.request.user.is_staff
    
    def delete(self, request, *args, **kwargs):
        categoria = self.get_object()
        messages.success(request, f"✅ Categoría '{categoria.nombre}' eliminada")
        return super().delete(request, *args, **kwargs)


# ==================== PLATILLOS ====================

class PlatilloListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Platillo
    template_name = 'roles/admin/Crud/platillos/platillos.html'
    context_object_name = 'platillos'
    paginate_by = 10
    
    def test_func(self):
        return self.request.user.is_staff
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_platillos'] = Platillo.objects.count()
        context['platillos_disponibles'] = Platillo.objects.filter(disponible=True).count()
        context['platillos_destacados'] = Platillo.objects.filter(destacado=True).count()
        context['categorias'] = CategoriaPlatillo.objects.filter(activo=True)
        return context


class PlatilloCreateView(CreateView):
    model = Platillo
    fields = ['nombre', 'descripcion', 'categoria', 'imagen_url', 'emojis', 'precio', 'disponible', 'destacado', 'orden']
    template_name = 'roles/admin/Crud/platillos/platillo_crear.html'
    success_url = reverse_lazy('platillos:lista')
    
    def test_func(self):
        return self.request.user.is_staff
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        categorias = CategoriaPlatillo.objects.filter(activo=True)
        context['categorias'] = categorias
        # Debug en consola
        print("=" * 50)
        print("DEBUG - Vista PlatilloCreateView")
        print(f"Categorías encontradas: {categorias.count()}")
        for c in categorias:
            print(f"  - ID: {c.id}, Nombre: {c.nombre}, Emoji: {c.emoji}")
        print("=" * 50)
        return context

    def form_valid(self, form):
        messages.success(self.request, f"✅ Platillo '{form.instance.nombre}' creado exitosamente")
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, "❌ Error al crear el platillo. Verifica los datos.")
        return super().form_invalid(form)


class PlatilloUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Platillo
    fields = ['nombre', 'descripcion', 'categoria', 'imagen_url', 'emojis', 'precio', 'disponible', 'destacado', 'orden']
    template_name = 'roles/admin/Crud/platillos/platillo_editar.html'
    success_url = reverse_lazy('platillos:lista')
    
    def test_func(self):
        return self.request.user.is_staff
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = CategoriaPlatillo.objects.filter(activo=True)
        return context

    def form_valid(self, form):
        messages.success(self.request, f"✅ Platillo '{form.instance.nombre}' actualizado exitosamente")
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, "❌ Error al actualizar el platillo")
        return super().form_invalid(form)


class PlatilloDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Platillo
    template_name = 'roles/admin/Crud/platillos/platillo_eliminar.html'
    success_url = reverse_lazy('platillos:lista')
    
    def test_func(self):
        return self.request.user.is_staff
    
    def delete(self, request, *args, **kwargs):
        platillo = self.get_object()
        messages.success(request, f"✅ Platillo '{platillo.nombre}' eliminado exitosamente")
        return super().delete(request, *args, **kwargs)

@login_required
@user_passes_test(es_admin)
def detalle_platillo(request, pk):
    platillo = get_object_or_404(Platillo, id=pk)
    return render(request, 'roles/admin/Crud/platillos/detalle_platillo.html', {'platillo': platillo})


# ==================== ESTADÍSTICAS ====================
@login_required
@user_passes_test(es_admin)
def estadisticas_platillos(request):
    """Vista de estadísticas de platillos"""
    platillos = Platillo.objects.all()
    categorias = CategoriaPlatillo.objects.all()
    hoy = datetime.now().date()
    
    # KPIs principales
    total_platillos = platillos.count()
    platillos_disponibles = platillos.filter(disponible=True).count()
    platillos_destacados = platillos.filter(destacado=True).count()
    precio_promedio = platillos.aggregate(Avg('precio'))['precio__avg'] or 0
    
    # Precio mínimo y máximo
    precio_minimo = platillos.aggregate(Min('precio'))['precio__min'] or 0
    precio_maximo = platillos.aggregate(Max('precio'))['precio__max'] or 0
    
    # Datos por categoría
    categorias_labels = []
    valores_categorias = []
    for cat in categorias:
        categorias_labels.append(f"{cat.emoji} {cat.nombre}")
        valores_categorias.append(platillos.filter(categoria=cat).count())
    
    # Evolución por mes (últimos 6 meses)
    meses = []
    cantidades = []
    
    for i in range(5, -1, -1):
        fecha = hoy - timedelta(days=30*i)
        mes_nombre = calendar.month_abbr[fecha.month]
        meses.append(mes_nombre)
        
        platillos_mes = platillos.filter(
            fecha_creacion__year=fecha.year,
            fecha_creacion__month=fecha.month
        ).count()
        cantidades.append(platillos_mes)
    
    context = {
        'total_platillos': total_platillos,
        'platillos_disponibles': platillos_disponibles,
        'platillos_destacados': platillos_destacados,
        'precio_promedio': precio_promedio,
        'precio_minimo': precio_minimo,
        'precio_maximo': precio_maximo,
        'categorias_labels': categorias_labels,
        'valores_categorias': valores_categorias,
        'meses': meses,
        'cantidades': cantidades,
    }
    
    return render(request, 'roles/admin/Crud/platillos/estadisticas_platillos.html', context)

@login_required
@user_passes_test(es_admin)
def export_estadisticas_platillos_pdf(request):
    """Exportar estadísticas de platillos a PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    
    platillos = Platillo.objects.all()
    categorias = CategoriaPlatillo.objects.all()
    hoy = datetime.now().date()
    
    total_platillos = platillos.count()
    platillos_disponibles = platillos.filter(disponible=True).count()
    platillos_destacados = platillos.filter(destacado=True).count()
    precio_promedio = platillos.aggregate(Avg('precio'))['precio__avg'] or 0
    precio_minimo = platillos.aggregate(Min('precio'))['precio__min'] or 0
    precio_maximo = platillos.aggregate(Max('precio'))['precio__max'] or 0
    
    categorias_labels = []
    valores_categorias = []
    for cat in categorias:
        categorias_labels.append(f"{cat.emoji} {cat.nombre}")
        valores_categorias.append(platillos.filter(categoria=cat).count())
    
    meses = []
    cantidades = []
    for i in range(5, -1, -1):
        fecha = hoy - timedelta(days=30*i)
        mes_nombre = calendar.month_name[fecha.month][:3]
        meses.append(mes_nombre)
        platillos_mes = platillos.filter(
            fecha_creacion__year=fecha.year,
            fecha_creacion__month=fecha.month
        ).count()
        cantidades.append(platillos_mes)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="estadisticas_platillos.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter),
                           leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
        fontSize=24, textColor=colors.HexColor('#d4af37'), alignment=TA_CENTER, spaceAfter=30)
    subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'],
        fontSize=12, textColor=colors.HexColor('#bba163'), alignment=TA_CENTER, spaceAfter=20)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'],
        fontSize=16, textColor=colors.HexColor('#f5d487'), spaceAfter=12, spaceBefore=20)
    
    story.append(Paragraph("LA FRAGATA GIRATORIA", title_style))
    story.append(Paragraph("Estadísticas de Platillos", subtitle_style))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # KPIs
    story.append(Paragraph("Métricas Principales", section_style))
    kpi_data = [
        ['Métrica', 'Valor'],
        ['Total Platillos', str(total_platillos)],
        ['Platillos Disponibles', str(platillos_disponibles)],
        ['Platillos Destacados', str(platillos_destacados)],
        ['Precio Promedio', f'${int(precio_promedio):,}'],
        ['Precio Mínimo', f'${int(precio_minimo):,}'],
        ['Precio Máximo', f'${int(precio_maximo):,}'],
    ]
    kpi_table = Table(kpi_data, colWidths=[150, 150])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4af37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 30))
    
    # Gráfico de distribución
    story.append(Paragraph("Distribución por Categoría", section_style))
    if sum(valores_categorias) > 0:
        drawing = Drawing(400, 220)
        pie = Pie()
        pie.x = 110
        pie.y = 30
        pie.width = 180
        pie.height = 180
        pie.data = valores_categorias
        pie.labels = categorias_labels
        pie.slices.strokeWidth = 0.5
        colores = ['#f5d487', '#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6']
        for i, color in enumerate(colores):
            if i < len(pie.slices):
                pie.slices[i].fillColor = colors.HexColor(color)
        drawing.add(pie)
        story.append(drawing)
    
    story.append(Spacer(1, 20))
    
    # Tabla por categoría
    categoria_data = [['Categoría', 'Cantidad', 'Porcentaje']]
    total = sum(valores_categorias)
    for i, label in enumerate(categorias_labels):
        porcentaje = (valores_categorias[i] / total * 100) if total > 0 else 0
        categoria_data.append([label, str(valores_categorias[i]), f'{porcentaje:.1f}%'])
    categoria_table = Table(categoria_data, colWidths=[150, 100, 100])
    categoria_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4af37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
    ]))
    story.append(categoria_table)
    story.append(Spacer(1, 30))
    
    # Evolución
    story.append(Paragraph("Evolución de Platillos por Mes", section_style))
    if max(cantidades) > 0:
        drawing2 = Drawing(500, 250)
        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 50
        bc.width = 400
        bc.height = 150
        bc.data = [cantidades]
        bc.categoryAxis.categoryNames = meses
        bc.valueAxis.valueMin = 0
        bc.bars[0].fillColor = colors.HexColor('#f5d487')
        drawing2.add(bc)
        story.append(drawing2)
    
    story.append(Spacer(1, 20))
    
    evolucion_data = [['Mes', 'Cantidad de Platillos']]
    for i, mes in enumerate(meses):
        evolucion_data.append([mes, str(cantidades[i])])
    evolucion_table = Table(evolucion_data, colWidths=[150, 150])
    evolucion_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4af37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
    ]))
    story.append(evolucion_table)
    story.append(Spacer(1, 30))
    
    story.append(Paragraph("Reporte generado automáticamente por el sistema La Fragata Giratoria", styles['Normal']))
    story.append(Paragraph("© 2025 - Todos los derechos reservados", styles['Normal']))
    
    doc.build(story)
    return response


# ==================== EXPORTACIONES ====================
@login_required
@user_passes_test(es_admin)
def export_platillos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Platillos"

    fondo_negro = PatternFill(start_color="0F0F0F", end_color="0F0F0F", fill_type="solid")
    dorado = "D4AF37"
    borde_dorado = Border(
        left=Side(style="thin", color=dorado),
        right=Side(style="thin", color=dorado),
        top=Side(style="thin", color=dorado),
        bottom=Side(style="thin", color=dorado),
    )

    headers = ["ID", "Emojis", "Nombre", "Categoría", "Precio", "Disponible", "Destacado", "Descripción"]
    ws.append(headers)

    for col in ws[1]:
        col.fill = fondo_negro
        col.font = Font(color=dorado, bold=True)
        col.alignment = Alignment(horizontal="center")
        col.border = borde_dorado

    for platillo in Platillo.objects.all():
        ws.append([
            platillo.id,
            platillo.emojis,
            platillo.nombre,
            platillo.categoria.nombre,
            float(platillo.precio),
            "Sí" if platillo.disponible else "No",
            "Sí" if platillo.destacado else "No",
            platillo.descripcion[:100]
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fondo_negro
            cell.font = Font(color="FFFFFF")
            cell.border = borde_dorado
            cell.alignment = Alignment(horizontal="center")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=platillos.xlsx"
    wb.save(response)
    return response

@login_required
@user_passes_test(es_admin)
def export_platillos_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=platillos.pdf"

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - 50

    # Título principal en dorado
    p.setFillColor(colors.HexColor("#D4AF37"))
    p.setFont("Helvetica-Bold", 20)
    p.drawString(200, y, "La Fragata Giratoria")
    y -= 30
    p.setFont("Helvetica-Bold", 16)
    p.drawString(220, y, "Listado de Platillos")
    y -= 30

    # Línea decorativa
    p.setStrokeColor(colors.HexColor("#D4AF37"))
    p.line(40, y, 550, y)
    y -= 20

    # Texto de los platillos en NEGRO
    p.setFont("Helvetica", 10)
    p.setFillColor(colors.black)  # Cambiado a negro

    for platillo in Platillo.objects.all():
        texto = f"{platillo.id} | {platillo.emojis} {platillo.nombre} | {platillo.categoria.nombre} | ${platillo.precio}"
        p.drawString(40, y, texto[:90])
        y -= 15

        if y < 40:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica", 10)
            p.setFillColor(colors.black)  # Mantener negro en nuevas páginas

    p.save()
    return response