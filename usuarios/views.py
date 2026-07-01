from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from datetime import datetime, timedelta
import calendar
from collections import Counter
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt
from django.core.mail import EmailMultiAlternatives, get_connection
import logging
from django.urls import reverse
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from usuarios.utils import registrar_actividad

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import json
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart

from .models import Usuario, Rol
from .forms import RegistroForm
from platillos.models import CategoriaPlatillo, Platillo
from pedidos.models import Pedido, PedidoItem, Cliente

logger = logging.getLogger(__name__)
from metodos_pago.models import MetodoPago

import datetime as dt

# ==================== FUNCIONES DE VERIFICACIÓN ====================
def es_admin(user):
    return (
        user.is_authenticated and
        getattr(user, 'estado', '').upper() == 'ACTIVO' and (
            user.is_staff or user.is_superuser or
            (hasattr(user, 'rol') and user.rol and getattr(user.rol, 'nombre_rol', '').upper() == 'ADMIN')
        )
    )

def es_cocinero(user):
    return (
        user.is_authenticated and
        getattr(user, 'estado', '').upper() == 'ACTIVO' and
        (hasattr(user, 'rol') and user.rol and getattr(user.rol, 'nombre_rol', '').upper() == 'COCINERO')
    )

def es_mesero(user):
    return (
        user.is_authenticated and
        getattr(user, 'estado', '').upper() == 'ACTIVO' and
        (hasattr(user, 'rol') and user.rol and getattr(user.rol, 'nombre_rol', '').upper() == 'MESERO')
    )

def es_cliente(user):
    return (
        user.is_authenticated and
        getattr(user, 'estado', '').upper() == 'ACTIVO' and
        (hasattr(user, 'rol') and user.rol and getattr(user.rol, 'nombre_rol', '').upper() == 'CLIENTE')
    )
    return user.is_authenticated and (user.rol and user.rol.nombre_rol == 'CLIENTE')

# ==================== LOGIN / LOGOUT / REGISTRO ====================
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard_redirect')

    if request.method == 'POST':
        email = request.POST.get('username')
        password = request.POST.get('password')

        usuario = Usuario.objects.filter(email=email).first()

        # CASO 1: Usuario no existe
        if not usuario:
            messages.error(request, "❌ Usuario o contraseña incorrectos")
            return render(request, 'home/login.html')

        # CASO 2: Usuario está INACTIVO
        if usuario.estado == 'INACTIVO':
            messages.warning(
                request,
                "⚠️ CUENTA INACTIVA: Tu cuenta está desactivada. Por favor, contacta con administración para activarla."
            )
            return render(request, 'home/login.html')

        # CASO 3: Usuario está SUSPENDIDO
        if usuario.estado == 'SUSPENDIDO':
            messages.error(
                request,
                "🚫 CUENTA SUSPENDIDA: Tu cuenta ha sido suspendida temporalmente. Contacta con administración para más información."
            )
            return render(request, 'home/login.html')

        # CASO 4: Usuario ACTIVO - intentar autenticación
        user = authenticate(request, username=email, password=password)

        if user is None:
            messages.error(request, "❌ Usuario o contraseña incorrectos")
            return render(request, 'home/login.html')

        # Login exitoso
        login(request, user)

        if es_admin(user):
            return redirect('/dashboard/admin/')
        elif es_cocinero(user):
            return redirect('/dashboard/cocinero/')
        elif es_mesero(user):
            return redirect('/mesero/pedidos/')
        else:
            return redirect('/dashboard/cliente/')

    return render(request, 'home/login.html')

def logout_view(request):
    logout(request)
    messages.success(request, "✅ Sesión cerrada correctamente")
    return redirect('login')

def registro_view(request):
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            nombre_usuario = form.cleaned_data['nombreUsuario']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            try:
                validate_password(password)
            except ValidationError as e:
                for error in e.messages:
                    messages.error(request, error)
                return render(request, 'home/registro.html', {'form': form})

            rol_cliente = Rol.objects.filter(nombre_rol='CLIENTE').first()
            
            if not rol_cliente:
                messages.error(request, "Error de configuración: rol CLIENTE no encontrado.")
                return render(request, 'home/registro.html', {'form': form})

            usuario = Usuario(
                nombre_usuario=nombre_usuario,
                email=email,
                estado='ACTIVO',
                rol=rol_cliente
            )
            usuario.set_password(password)
            usuario.save()
            registrar_actividad(usuario, 'usuario', f"Nuevo usuario registrado: {usuario.nombre_usuario}")
            
            # ✅ Hacer login automático y redirigir al dashboard del cliente
            login(request, usuario)
            messages.success(request, "✅ Registro exitoso. ¡Bienvenido!")
            return redirect('cliente_dashboard')
    else:
        form = RegistroForm()

    return render(request, 'home/registro.html', {'form': form})


# ==================== RECUPERACIÓN DE CONTRASEÑA ====================

def _construir_enlace_recuperacion(request, usuario):
    uid = urlsafe_base64_encode(force_bytes(usuario.pk))
    token = default_token_generator.make_token(usuario)
    path = reverse('restablecer_contraseña', kwargs={'uidb64': uid, 'token': token})
    if request:
        return request.build_absolute_uri(path)
    return f"{settings.SITE_URL.rstrip('/')}{path}"


@ensure_csrf_cookie
def solicitar_recuperacion_contraseña(request):
    """Envía enlace de recuperación al correo del usuario."""

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()

        if not email:
            messages.error(request, "Por favor ingresa tu correo electrónico.")
            return render(request, 'home/recuperar_contraseña.html')

        try:
            usuario = Usuario.objects.get(email__iexact=email)
            reset_link = _construir_enlace_recuperacion(request, usuario)

            asunto = "La Fragata Giratoria - Recupera tu contraseña"

            mensaje_html = f"""
            <html>
            <body style="font-family: Arial, sans-serif; background-color: #0f0f0f; color: #ffffff; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; background-color: #1a1a1a;
                            border-left: 4px solid #d4af37; padding: 30px; border-radius: 8px;">
                    <h1 style="color: #d4af37; text-align: center;">La Fragata Giratoria</h1>
                    <p>Hola <strong>{usuario.nombre_usuario}</strong>,</p>

                    <p style="color: #bba163;">
                        Recibimos una solicitud para restablecer tu contraseña.
                        Haz clic en el botón para continuar:
                    </p>

                    <p style="text-align: center; margin: 30px 0;">
                        <a href="{reset_link}"
                           style="background-color: #d4af37; color: #0f0f0f; padding: 12px 30px;
                                  text-decoration: none; border-radius: 4px; font-weight: bold;">
                            Restablecer contraseña
                        </a>
                    </p>

                    <p style="font-size: 12px; color: #999;">
                        Si no solicitaste esto, ignora este correo. El enlace expira en 24 horas.
                    </p>
                </div>
            </body>
            </html>
            """

            mensaje_texto = (
                f"Hola {usuario.nombre_usuario},\n\n"
                f"Para restablecer tu contraseña abre este enlace:\n{reset_link}\n\n"
                "Si no solicitaste esto, ignora este mensaje."
            )

            import os
            import resend

            resend.api_key = os.getenv("RESEND_API_KEY")

            resend.Emails.send({
                "from": "La Fragata Giratoria <onboarding@resend.dev>",
                "to": [usuario.email],
                "subject": asunto,
                "html": mensaje_html,
                "text": mensaje_texto
            })

            messages.success(
                request,
                f"Se envió un correo a {email}. Revisa tu bandeja de entrada y spam."
            )
            return redirect('login')

        except Usuario.DoesNotExist:
            messages.info(
                request,
                "Si el correo existe en nuestro sistema, recibirás instrucciones para recuperar tu contraseña.",
            )
            return redirect('login')

        except Exception as exc:
            logger.exception('Error al enviar correo de recuperación de contraseña')
            messages.error(
                request,
                f'No se pudo enviar el correo. Error: {exc}'
            )
            return render(request, 'home/recuperar_contraseña.html')

    return render(request, 'home/recuperar_contraseña.html')

print("=" * 50)
print("EMAIL_BACKEND:", settings.EMAIL_BACKEND)
print("EMAIL_HOST:", settings.EMAIL_HOST)
print("EMAIL_PORT:", settings.EMAIL_PORT)
print("EMAIL_HOST_USER:", settings.EMAIL_HOST_USER)
print("EMAIL_USE_TLS:", settings.EMAIL_USE_TLS)
print("EMAIL_USE_SSL:", settings.EMAIL_USE_SSL)
print("=" * 50)


@ensure_csrf_cookie
def restablecer_contraseña(request, uidb64, token):
    """Formulario para establecer nueva contraseña desde el enlace del correo."""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        usuario = Usuario.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, Usuario.DoesNotExist):
        usuario = None

    if usuario is None or not default_token_generator.check_token(usuario, token):
        messages.error(request, "El enlace no es válido o ya expiró. Solicita uno nuevo.")
        return redirect('password_reset')

    if request.method == 'POST':
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        if not password1 or len(password1) < 8:
            messages.error(request, "La contraseña debe tener al menos 8 caracteres.")
        elif password1 != password2:
            messages.error(request, "Las contraseñas no coinciden.")
        else:
            try:
                validate_password(password1, usuario)
                usuario.set_password(password1)
                usuario.save()
                messages.success(request, "Contraseña actualizada. Ya puedes iniciar sesión.")
                return redirect('login')
            except ValidationError as e:
                for err in e.messages:
                    messages.error(request, err)

    return render(request, 'home/restablecer_contraseña.html', {'validlink': True})

# ==================== DASHBOARD REDIRECT ====================
def dashboard_redirect(request):
    if not request.user.is_authenticated:
        return redirect('login')

    if es_admin(request.user):
        return redirect('/dashboard/admin/')
    elif es_cocinero(request.user):
        return redirect('/dashboard/cocinero/')
    elif es_mesero(request.user):
        return redirect('/mesero/pedidos/')
    else:
        return redirect('/dashboard/cliente/')

# ==================== DASHBOARD ADMIN ====================
@login_required
def dashboard_admin(request):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')

    context = {
        'total_usuarios': Usuario.objects.count(),
        'total_usuarios_activos': Usuario.objects.filter(estado='ACTIVO').count(),
        'total_usuarios_inactivos': Usuario.objects.filter(estado='INACTIVO').count(),
    }
    return render(request, 'roles/admin/dashboard.html', context)

# ==================== DASHBOARD COCINERO ====================
@login_required
def dashboard_cocinero(request):
    if not es_cocinero(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    return redirect('cocina_pedidos')

@login_required
def cocina_pedidos(request):
    if not es_cocinero(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')

    pedidos_pendientes = Pedido.objects.filter(estado='PENDIENTE').order_by('-fecha')
    pedidos_en_proceso = Pedido.objects.filter(estado='EN PROCESO').order_by('-fecha')
    
    context = {
        'pedidos_pendientes': pedidos_pendientes,
        'pedidos_en_proceso': pedidos_en_proceso,
    }
    return render(request, 'roles/Cocinero/cocina_dashboard.html', context)

@login_required
def cocinero_actualizar_estado(request, pedido_id):
    if not es_cocinero(request.user):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')

    if request.method == 'POST':
        try:
            pedido = get_object_or_404(Pedido, id_pedido=pedido_id)
            nuevo_estado = request.POST.get('estado')
            
            if not nuevo_estado:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Estado no proporcionado'})
                messages.error(request, "Estado no proporcionado")
                return redirect('cocina_pedidos')
            
            pedido.estado = nuevo_estado
            pedido.save()
            
            mensaje = "Pedido enviado correctamente"
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': mensaje,
                    'nuevo_estado': nuevo_estado
                })
            else:
                messages.success(request, f"✅ {mensaje}")
                return redirect('cocina_pedidos')
                
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)}, status=500)
            messages.error(request, f"Error: {str(e)}")
            return redirect('cocina_pedidos')

    return redirect('cocina_pedidos')

@login_required
def cocina_check_pedidos(request):
    if not es_cocinero(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    
    pedidos = Pedido.objects.filter(estado__in=['PENDIENTE', 'EN PROCESO']).order_by('-fecha')
    
    pedidos_data = []
    for pedido in pedidos:
        platillos_nombres = [item.nombre_platillo for item in pedido.items.all()]
        platillos_str = ', '.join(platillos_nombres[:3])
        if len(platillos_nombres) > 3:
            platillos_str += f' y {len(platillos_nombres) - 3} más'
        
        pedidos_data.append({
            'id': pedido.id_pedido,
            'platillos': platillos_str
        })
    
    return JsonResponse({'new_pedidos': pedidos_data})

# ==================== DASHBOARD MESERO ====================
@login_required
def dashboard_mesero(request):
    if not es_mesero(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    return redirect('mesero_pedidos')

@login_required
def mesero_pedidos(request):
    if not es_mesero(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')

    pedidos = Pedido.objects.filter(estado='COMPLETADO').order_by('-fecha')
    return render(request, 'roles/Mesero/mesero_dashboard.html', {'pedidos': pedidos})

@login_required
def mesero_entregar_pedido(request, pedido_id):
    if not es_mesero(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')

    if request.method == 'POST':
        try:
            pedido = get_object_or_404(Pedido, id_pedido=pedido_id)
            pedido.estado = 'ENTREGADO'
            pedido.estado_cocina = 'ENTREGADO'
            pedido.estado_mesero = 'ENTREGADO'
            pedido.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'✅ Pedido #{pedido.id_pedido} entregado al cliente'
                })
            else:
                messages.success(request, f"✅ Pedido #{pedido.id_pedido} entregado al cliente")
                return redirect('mesero_pedidos')
                
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)}, status=500)
            else:
                messages.error(request, f"Error al entregar el pedido: {str(e)}")
                return redirect('mesero_pedidos')

    return redirect('mesero_pedidos')

@login_required
def mesero_check_pedidos(request):
    if not es_mesero(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    
    pedidos = Pedido.objects.filter(estado='COMPLETADO').order_by('-fecha')
    pedidos_data = [{'id': p.id_pedido} for p in pedidos]
    
    return JsonResponse({'new_pedidos': pedidos_data})

# ==================== DASHBOARD CLIENTE ====================
@login_required
def dashboard_cliente(request):
    if not es_cliente(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    return render(request, 'roles/Cliente/dashboard.html')

@login_required
def cliente_inicio(request):
    if not es_cliente(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    return render(request, 'roles/Cliente/dashboard.html')

@login_required
def cliente_menu(request):
    if not es_cliente(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')

    categorias = CategoriaPlatillo.objects.filter(activo=True).prefetch_related('platillos')
    for categoria in categorias:
        categoria.platillos_disponibles = categoria.platillos.filter(disponible=True)

    return render(request, 'roles/Cliente/menu.html', {'categorias': categorias})

@login_required
def cliente_carrito(request):
    if not es_cliente(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')

    carrito = request.session.get('carrito', {})
    total = sum(item['precio'] * item['cantidad'] for item in carrito.values())
    
    # LOG para depurar
    print(f"=== CARRITO DEBUG ===")
    print(f"Usuario: {request.user.nombre_usuario}")
    print(f"Carrito: {carrito}")
    print(f"Total: {total}")
    
    # Verificar si hay imágenes
    for key, item in carrito.items():
        print(f"Item {key}: {item.get('nombre')} - imagen: {item.get('imagen_url', 'NO IMAGEN')}")

    return render(request, 'roles/Cliente/carrito.html', {
        'carrito': carrito,
        'total': total
    })
@login_required
@csrf_exempt
def cliente_carrito_agregar(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})

    id_platillo = request.POST.get('idPlatillo')
    cantidad = int(request.POST.get('cantidad', 1))

    if not id_platillo:
        return JsonResponse({'success': False, 'error': 'ID de platillo requerido'})

    try:
        platillo = Platillo.objects.get(id=id_platillo)

        carrito = request.session.get('carrito', {})
        str_id = str(id_platillo)

        # ✅ Esto es CLAVE - Normaliza la URL de la imagen
        normalized_imagen_url = platillo.imagen_url.strip() if platillo.imagen_url else ''
        if normalized_imagen_url and not normalized_imagen_url.startswith(('http://', 'https://')):
            normalized_imagen_url = platillo.get_imagen_static_path()

        if str_id in carrito:
            carrito[str_id]['cantidad'] += cantidad
        else:
            carrito[str_id] = {
                'id': platillo.id,
                'nombre': platillo.nombre,
                'descripcion': platillo.descripcion,
                'precio': float(platillo.precio),
                'cantidad': cantidad,
                'emojis': platillo.emojis,
                'imagen_url': normalized_imagen_url,  # ✅ Aquí se guarda la imagen
            }

        request.session['carrito'] = carrito
        request.session.modified = True

        return JsonResponse({
            'success': True,
            'message': f'{platillo.nombre} agregado al carrito',
            'carrito_count': sum(item['cantidad'] for item in carrito.values())
        })

    except Platillo.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Platillo no encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def cliente_registrar_pedido(request):
    """Registrar el pedido del cliente"""
    if request.method != 'POST':
        return redirect('cliente_carrito')

    carrito = request.session.get('carrito', {})

    if not carrito:
        messages.error(request, "No hay productos en el carrito")
        return redirect('cliente_menu')

    id_metodo_pago = request.POST.get('idMetodoPago')
    if not id_metodo_pago:
        messages.error(request, "Seleccione un método de pago")
        return redirect('cliente_carrito')

    try:
        from pedidos.models import Pedido, PedidoItem, Cliente
        from metodos_pago.models import MetodoPago

        cliente, created = Cliente.objects.get_or_create(
            email=request.user.email,
            defaults={
                'nombre': request.user.nombre_usuario,
                'telefono': '',
                'tipo': 'LOCAL'
            }
        )

        metodo_pago = MetodoPago.objects.get(id_metodo_pago=id_metodo_pago)

        total_pedido = 0
        for key, item in carrito.items():
            total_pedido += float(item['precio']) * int(item['cantidad'])

        pedido = Pedido.objects.create(
            id_cliente=cliente,
            id_usuario=request.user,
            id_metodo_pago=metodo_pago,
            fecha=timezone.now().date(),
            estado='PENDIENTE',
            estado_cocina='PENDIENTE',
            total=total_pedido
        )

        for key, item in carrito.items():
            PedidoItem.objects.create(
                pedido=pedido,
                nombre_platillo=item['nombre'],
                cantidad=int(item['cantidad']),
                precio_unitario=float(item['precio']),
                subtotal=float(item['precio']) * int(item['cantidad'])
            )

        # Limpiar carrito
        request.session['carrito'] = {}
        request.session.modified = True

        # Devolver respuesta JSON correcta
        return JsonResponse({
            'success': True,
            'pedido_id': pedido.id_pedido,
            'message': f'Pedido #{pedido.id_pedido} enviado a la cocina'
        })

    except Exception as e:
        print(f"Error: {e}")
        return JsonResponse({
            'success': False, 
            'error': str(e)
        })
        
# ==================== LOGIN / LOGOUT / REGISTRO ====================
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard_redirect')

    if request.method == 'POST':
        identifier = request.POST.get('email') or request.POST.get('username') or request.POST.get('identifier')
        password = request.POST.get('password')

        if not identifier or not password:
            messages.error(request, "Ingrese usuario y contraseña.")
            return render(request, 'home/login.html')

        # Buscar usuario primero (sin autenticar) para mostrar el aviso de estado
        usuario = Usuario.objects.filter(
            Q(email__iexact=identifier) | Q(nombre_usuario__iexact=identifier)
        ).first()

        if usuario and getattr(usuario, 'estado', None):
            estado = usuario.estado.upper()
            if estado == 'SUSPENDIDO':
                messages.error(request, "⚠️ Cuenta suspendida. Contacte al administrador.")
                return render(request, 'home/login.html')
            if estado == 'INACTIVO':
                messages.error(request, "⚠️ Cuenta inactiva. Verifique su correo o contacte soporte.")
                return render(request, 'home/login.html')

        # Si no hay usuario con estado bloqueante, proceder a autenticar
        user = authenticate(request, username=identifier, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard_redirect')

        messages.error(request, "Nombre de usuario o contraseña incorrectos.")

    return render(request, 'home/login.html')

# ==================== CRUD USUARIOS (SOLO ADMIN) ====================
@login_required
def lista_usuarios(request):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    
    usuarios = Usuario.objects.all().select_related('rol')
    context = {
        'usuarios': usuarios,
        'total_usuarios': usuarios.count(),
        'usuarios_activos': usuarios.filter(estado='ACTIVO').count(),
        'usuarios_inactivos': usuarios.filter(estado='INACTIVO').count(),
        'usuarios_suspendidos': usuarios.filter(estado='SUSPENDIDO').count(),
    }
    return render(request, 'roles/admin/Crud/usuarios/usuarios.html', context)

@login_required
def estadisticas_usuarios(request):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    
    usuarios = Usuario.objects.all()
    context = {
        'total_usuarios': usuarios.count(),
        'usuarios_activos': usuarios.filter(estado='ACTIVO').count(),
        'usuarios_inactivos': usuarios.filter(estado='INACTIVO').count(),
        'usuarios_suspendidos': usuarios.filter(estado='SUSPENDIDO').count(),
        'usuarios_por_rol': {
            'admin': usuarios.filter(rol__nombre_rol='ADMIN').count(),
            'cocinero': usuarios.filter(rol__nombre_rol='COCINERO').count(),
            'mesero': usuarios.filter(rol__nombre_rol='MESERO').count(),
            'cliente': usuarios.filter(rol__nombre_rol='CLIENTE').count(),
        }
    }
    return render(request, 'roles/admin/Crud/usuarios/estadisticas_usuarios.html', context)

@login_required
def crear_usuario(request):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    
    if request.method == 'POST':
        nombre_usuario = request.POST.get('nombre_usuario')
        email = request.POST.get('email')
        password = request.POST.get('password')
        rol_id = request.POST.get('rol_id')
        estado = request.POST.get('estado', 'ACTIVO')

        if nombre_usuario and email and password and rol_id:
            if Usuario.objects.filter(email=email).exists():
                messages.error(request, "❌ El email ya está registrado")
            elif Usuario.objects.filter(nombre_usuario=nombre_usuario).exists():
                messages.error(request, "❌ El nombre de usuario ya está registrado")
            else:
                rol = Rol.objects.filter(id_rol=rol_id).first()
                if not rol:
                    messages.error(request, "❌ El rol seleccionado no existe")
                else:
                    usuario = Usuario(
                        nombre_usuario=nombre_usuario,
                        email=email,
                        estado=estado,
                        rol=rol
                    )
                    usuario.set_password(password)
                    usuario.save()
                    registrar_actividad(request.user, 'usuario', f"Usuario '{nombre_usuario}' creado")
                    messages.success(request, f"✅ Usuario '{nombre_usuario}' creado exitosamente")
                    return redirect('usuarios:lista')
        else:
            messages.error(request, "❌ Todos los campos son obligatorios")

    roles = Rol.objects.all()
    return render(request, 'roles/admin/Crud/usuarios/crear_usuario.html', {'roles': roles})

@login_required
def editar_usuario(request, pk):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    
    usuario = get_object_or_404(Usuario, id_usuario=pk)
    
    if request.method == 'POST':
        nuevo_nombre = request.POST.get('nombre_usuario')
        nuevo_email = request.POST.get('email')
        nuevo_estado = request.POST.get('estado')
        nuevo_rol_id = request.POST.get('rol_id')
        nueva_password = request.POST.get('nueva_password')
        
        if Usuario.objects.filter(email=nuevo_email).exclude(id_usuario=pk).exists():
            messages.error(request, "❌ El email ya está registrado por otro usuario.")
        elif Usuario.objects.filter(nombre_usuario=nuevo_nombre).exclude(id_usuario=pk).exists():
            messages.error(request, "❌ El nombre de usuario ya está en uso.")
        else:
            rol = Rol.objects.filter(id_rol=nuevo_rol_id).first()
            if not rol:
                messages.error(request, "❌ El rol seleccionado no es válido.")
            else:
                usuario.nombre_usuario = nuevo_nombre
                usuario.email = nuevo_email
                usuario.estado = nuevo_estado
                usuario.rol = rol
                
                if nueva_password:
                    usuario.set_password(nueva_password)
                
                usuario.save()
                registrar_actividad(request.user, 'editar', f"Usuario '{usuario.nombre_usuario}' actualizado")
                messages.success(request, f"✅ Usuario '{usuario.nombre_usuario}' actualizado")
                return redirect('usuarios:lista')
    
    roles = Rol.objects.all()
    context = {
        'usuario': usuario,
        'roles': roles
    }
    return render(request, 'roles/admin/Crud/usuarios/editar_usuario.html', context)

@login_required
def eliminar_usuario(request, pk):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    
    usuario = get_object_or_404(Usuario, id_usuario=pk)
    if request.method == 'POST':
        nombre = usuario.nombre_usuario
        registrar_actividad(request.user, 'eliminar', f"Usuario '{nombre}' eliminado")
        usuario.delete()
        messages.success(request, f"✅ Usuario '{nombre}' eliminado", extra_tags='delete')
        return redirect('usuarios:lista')
    
    return render(request, 'roles/admin/Crud/usuarios/eliminar_usuario.html', {'usuario': usuario})

@login_required
def detalle_usuario(request, pk):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    
    usuario = get_object_or_404(Usuario.objects.select_related('rol'), id_usuario=pk)
    return render(request, 'roles/admin/Crud/usuarios/detalle_usuario.html', {'usuario': usuario})

# ==================== EXPORTACIONES ====================
@login_required
def export_usuarios_excel(request):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    fondo_negro = PatternFill(start_color="0F0F0F", end_color="0F0F0F", fill_type="solid")
    dorado = "D4AF37"
    borde_dorado = Border(
        left=Side(style="thin", color=dorado),
        right=Side(style="thin", color=dorado),
        top=Side(style="thin", color=dorado),
        bottom=Side(style="thin", color=dorado),
    )

    headers = ["ID", "Usuario", "Email", "Rol", "Estado", "Fecha Creación"]
    ws.append(headers)

    for col in ws[1]:
        col.fill = fondo_negro
        col.font = Font(color=dorado, bold=True)
        col.alignment = Alignment(horizontal="center")
        col.border = borde_dorado

    usuarios = Usuario.objects.all().select_related('rol')
    for u in usuarios:
        ws.append([
            u.id_usuario,
            u.nombre_usuario,
            u.email,
            u.rol.nombre_rol if u.rol else 'Sin rol',
            u.estado,
            u.fecha_creacion.strftime('%d/%m/%Y %H:%M') if u.fecha_creacion else ''
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fondo_negro
            cell.font = Font(color="FFFFFF")
            cell.border = borde_dorado
            cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    wb.save(response)
    return response

@login_required
def export_usuarios_pdf(request):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=usuarios.pdf'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - 50

    p.setFillColor(colors.HexColor("#D4AF37"))
    p.setFont("Helvetica-Bold", 20)
    p.drawString(200, y, "La Fragata Giratoria")
    y -= 30

    p.setFont("Helvetica-Bold", 16)
    p.drawString(220, y, "Listado de Usuarios")
    y -= 30

    p.setStrokeColor(colors.HexColor("#D4AF37"))
    p.line(40, y, 550, y)
    y -= 20

    usuarios = Usuario.objects.all().select_related('rol')
    p.setFont("Helvetica", 10)

    for u in usuarios:
        texto = f"{u.id_usuario} | {u.nombre_usuario} | {u.email} | {u.rol.nombre_rol if u.rol else ''} | {u.estado}"
        p.drawString(40, y, texto[:80])
        y -= 15

        if y < 40:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica", 10)

    p.save()
    return response

@login_required
def export_estadisticas_usuarios_pdf(request):
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')
    
    usuarios = Usuario.objects.all()

    total_usuarios = usuarios.count()
    usuarios_activos = usuarios.filter(estado='ACTIVO').count()
    usuarios_inactivos = usuarios.filter(estado='INACTIVO').count()
    usuarios_suspendidos = usuarios.filter(estado='SUSPENDIDO').count()

    usuarios_admin = usuarios.filter(rol__nombre_rol='ADMIN').count()
    usuarios_cocinero = usuarios.filter(rol__nombre_rol='COCINERO').count()
    usuarios_mesero = usuarios.filter(rol__nombre_rol='MESERO').count()
    usuarios_cliente = usuarios.filter(rol__nombre_rol='CLIENTE').count()

    estados_labels = ['Activos', 'Inactivos', 'Suspendidos']
    estados_data = [usuarios_activos, usuarios_inactivos, usuarios_suspendidos]

    roles_labels = ['Administradores', 'Cocineros', 'Meseros', 'Clientes']
    roles_data = [usuarios_admin, usuarios_cocinero, usuarios_mesero, usuarios_cliente]

    porcentaje_activos = (usuarios_activos / total_usuarios * 100) if total_usuarios > 0 else 0
    porcentaje_inactivos = (usuarios_inactivos / total_usuarios * 100) if total_usuarios > 0 else 0
    porcentaje_suspendidos = (usuarios_suspendidos / total_usuarios * 100) if total_usuarios > 0 else 0

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="estadisticas_usuarios.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter),
                            leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                  fontSize=24, textColor=colors.HexColor('#d4af37'),
                                  alignment=TA_CENTER, spaceAfter=30)
    subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'],
                                     fontSize=12, textColor=colors.HexColor('#bba163'),
                                     alignment=TA_CENTER, spaceAfter=20)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'],
                                    fontSize=16, textColor=colors.HexColor('#f5d487'),
                                    spaceAfter=12, spaceBefore=20)

    story.append(Paragraph("LA FRAGATA GIRATORIA", title_style))
    story.append(Paragraph("Estadísticas de Usuarios", subtitle_style))
    story.append(Spacer(1, 20))

    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph(f"Fecha de generación: {fecha_actual}", styles['Normal']))
    story.append(Spacer(1, 20))

    story.append(Paragraph("Métricas Principales", section_style))

    kpi_data = [
        ['Métrica', 'Valor', 'Porcentaje'],
        ['Total Usuarios', str(total_usuarios), '100%'],
        ['Usuarios Activos', str(usuarios_activos), f'{porcentaje_activos:.1f}%'],
        ['Usuarios Inactivos', str(usuarios_inactivos), f'{porcentaje_inactivos:.1f}%'],
        ['Usuarios Suspendidos', str(usuarios_suspendidos), f'{porcentaje_suspendidos:.1f}%'],
    ]
    kpi_table = Table(kpi_data, colWidths=[150, 120, 120])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4af37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 30))

    story.append(Paragraph("Distribución por Estado", section_style))
    if sum(estados_data) > 0:
        drawing = Drawing(400, 220)
        pie = Pie()
        pie.x = 110
        pie.y = 30
        pie.width = 180
        pie.height = 180
        pie.data = estados_data
        pie.labels = estados_labels
        pie.slices.strokeWidth = 0.5
        for i, color in enumerate(['#10b981', '#f59e0b', '#ef4444']):
            if i < len(pie.slices):
                pie.slices[i].fillColor = colors.HexColor(color)
        drawing.add(pie)
        story.append(drawing)
    story.append(Spacer(1, 20))

    estados_table_data = [['Estado', 'Cantidad', 'Porcentaje']]
    for label, value in zip(estados_labels, estados_data):
        porcentaje = (value / total_usuarios * 100) if total_usuarios > 0 else 0
        estados_table_data.append([label, str(value), f'{porcentaje:.1f}%'])
    estados_table = Table(estados_table_data, colWidths=[150, 100, 100])
    estados_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4af37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
    ]))
    story.append(estados_table)
    story.append(Spacer(1, 30))

    story.append(Paragraph("Distribución por Rol", section_style))
    if max(roles_data) > 0:
        drawing2 = Drawing(500, 250)
        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 50
        bc.width = 400
        bc.height = 150
        bc.data = [roles_data]
        bc.categoryAxis.categoryNames = roles_labels
        bc.valueAxis.valueMin = 0
        bc.valueAxis.valueMax = max(roles_data) * 1.2 if max(roles_data) > 0 else 10
        bc.bars[0].fillColor = colors.HexColor('#f5d487')
        bc.bars[0].strokeColor = colors.HexColor('#d4af37')
        drawing2.add(bc)
        story.append(drawing2)
    story.append(Spacer(1, 20))

    roles_table_data = [['Rol', 'Cantidad', 'Porcentaje']]
    for label, value in zip(roles_labels, roles_data):
        porcentaje = (value / total_usuarios * 100) if total_usuarios > 0 else 0
        roles_table_data.append([label, str(value), f'{porcentaje:.1f}%'])
    roles_table = Table(roles_table_data, colWidths=[150, 100, 100])
    roles_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4af37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
    ]))
    story.append(roles_table)
    story.append(Spacer(1, 30))

    story.append(Paragraph("Información Adicional", section_style))
    info_data = [
        ['Usuarios Registrados', f'{total_usuarios} usuarios'],
        ['Usuarios Activos', f'{usuarios_activos} ({porcentaje_activos:.1f}%)'],
        ['Usuarios Inactivos', f'{usuarios_inactivos} ({porcentaje_inactivos:.1f}%)'],
        ['Usuarios Suspendidos', f'{usuarios_suspendidos} ({porcentaje_suspendidos:.1f}%)'],
        ['Rol más común', max(zip(roles_labels, roles_data), key=lambda x: x[1])[0] if any(roles_data) else 'N/A'],
    ]
    info_table = Table(info_data, colWidths=[150, 250])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d4af37')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 30))

    story.append(Paragraph("Reporte generado automáticamente por el sistema La Fragata Giratoria", styles['Normal']))
    story.append(Paragraph("© 2025 - Todos los derechos reservados", styles['Normal']))

    doc.build(story)
    return response


@login_required
def eliminar_multiple_usuarios(request):
    """Eliminar múltiples usuarios vía POST (form o JSON). Requiere admin."""
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')

    if request.method != 'POST':
        messages.error(request, 'Método no permitido')
        return redirect('usuarios:lista')

    ids = []
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
        ids = payload.get('ids') or []
    except Exception:
        ids = []

    if not ids:
        ids = request.POST.getlist('ids') or request.POST.get('delete_ids', '')

    if isinstance(ids, str):
        ids = [i for i in ids.split(',') if i.strip()]

    if not ids:
        messages.error(request, 'No se recibieron IDs para eliminar')
        return redirect('usuarios:lista')

    qs = Usuario.objects.filter(id_usuario__in=ids)
    deleted = qs.count()
    qs.delete()
    messages.success(request, f'✅ {deleted} usuario(s) eliminado(s)')
    return redirect('usuarios:lista')


@login_required
def exportar_usuarios_seleccionados(request):
    """Exportar usuarios seleccionados a Excel (GET ?ids=1,2 o POST ids[])."""
    if not es_admin(request.user):
        messages.error(request, "No tienes permiso para acceder a esta sección")
        return redirect('dashboard_redirect')

    ids = request.GET.get('ids', '')
    if not ids and request.method == 'POST':
        ids = request.POST.getlist('ids') or request.POST.get('ids', '')

    if isinstance(ids, str):
        ids_list = [i for i in ids.split(',') if i.strip()]
    else:
        ids_list = ids

    if ids_list:
        usuarios_qs = Usuario.objects.filter(id_usuario__in=ids_list)
    else:
        usuarios_qs = Usuario.objects.none()

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios Seleccionados"

    headers = ["ID", "Usuario", "Email", "Rol", "Estado", "Fecha Creación"]
    ws.append(headers)

    for u in usuarios_qs:
        ws.append([
            u.id_usuario,
            u.nombre_usuario,
            u.email,
            u.rol.nombre_rol if u.rol else 'Sin rol',
            u.estado,
            u.fecha_creacion.strftime('%d/%m/%Y %H:%M') if u.fecha_creacion else ''
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=usuarios_seleccionados.xlsx'
    wb.save(response)
    return response