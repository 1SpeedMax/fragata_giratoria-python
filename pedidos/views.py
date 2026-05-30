import io
import json
from datetime import datetime, date
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Sum, Avg, Count
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- IMPORTACIÓN DE MODELOS ---
from .models import Pedido, PedidoItem, Cliente
from productos.models import Producto
from platillos.models import Platillo
from metodos_pago.models import MetodoPago
from usuarios.models import Usuario

# --- REPORTLAB PARA PDF ---
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors

from django.views.generic import ListView
from .models import Pedido

# ==================== VISTAS DE PEDIDOS ====================

def lista_pedidos(request):
    pedidos = Pedido.objects.all().order_by('-fecha') 
    return render(request, 'roles/admin/Crud/pedidos/pedidos.html', {'pedidos': pedidos})

def crear_pedido(request):
    if request.method == 'POST':
        # Obtener datos del formulario
        fecha = request.POST.get('fecha')
        id_cliente = request.POST.get('id_cliente')
        id_usuario = request.POST.get('id_usuario')
        id_metodo_pago = request.POST.get('id_metodo_pago')
        observaciones = request.POST.get('observaciones', '')
        
        try:
            # Procesar relaciones opcionales
            cliente = None
            if id_cliente:
                cliente = Cliente.objects.get(id_cliente=id_cliente)
            
            usuario = None
            if id_usuario:
                usuario = Usuario.objects.get(id=id_usuario)
            
            metodo_pago = None
            if id_metodo_pago:
                metodo_pago = MetodoPago.objects.get(id_metodo_pago=id_metodo_pago)
            
            # Crear pedido
            total_pedido = Decimal('0')
            
            pedido = Pedido.objects.create(
                fecha=fecha or date.today(),
                id_cliente=cliente,
                id_usuario=usuario,
                id_metodo_pago=metodo_pago,
                estado='PENDIENTE',
                observaciones=observaciones,
                total=total_pedido
            )
            
            # Procesar items
            items_data = request.POST.getlist('items[]')
            total_pedido = Decimal('0')
            
            for item_json in items_data:
                if item_json:
                    try:
                        item = json.loads(item_json)
                        platillo_id = item.get('platillo_id')
                        cantidad = int(item.get('cantidad', 1))
                        precio_unitario = Decimal(str(item.get('precio', 0)))
                        subtotal = precio_unitario * cantidad
                        
                        if platillo_id:
                            platillo = Platillo.objects.get(id=platillo_id)
                            PedidoItem.objects.create(
                                pedido=pedido,
                                platillo=platillo,
                                nombre_platillo=item.get('nombre', platillo.nombre),
                                cantidad=cantidad,
                                precio_unitario=precio_unitario,
                                subtotal=subtotal
                            )
                            total_pedido += subtotal
                    except (json.JSONDecodeError, ValueError, Platillo.DoesNotExist):
                        continue
            
            # Actualizar total del pedido
            pedido.total = total_pedido
            pedido.save()
            
            messages.success(request, f"✅ Pedido #{pedido.id_pedido} creado exitosamente con total: ${total_pedido}")
            return redirect('pedidos:lista')
        
        except (Cliente.DoesNotExist, User.DoesNotExist, MetodoPago.DoesNotExist) as e:
            messages.error(request, f"Error al crear pedido: {str(e)}")
            return redirect('pedidos:nuevo')
    
    # GET - mostrar formulario
    platillos = Platillo.objects.filter(disponible=True)
    clientes = Cliente.objects.all()
    usuarios = Usuario.objects.all()
    metodos_pago = MetodoPago.objects.all()
    
    # Convertir platillos a JSON para JavaScript
    platillos_json = json.dumps([
        {
            'id': p.id,
            'nombre': p.nombre,
            'precio': float(p.precio),
            'descripcion': p.descripcion
        }
        for p in platillos
    ])
    
    context = {
        'platillos': platillos,
        'platillos_json': platillos_json,
        'clientes': clientes,
        'usuarios': usuarios,
        'metodos_pago': metodos_pago,
        'hoy': date.today()
    }
    
    return render(request, 'roles/admin/Crud/pedidos/pedidos_crear.html', context)

def detalle_pedido(request, id_pedido):
    pedido = get_object_or_404(Pedido, id_pedido=id_pedido)
    return render(request, 'roles/admin/Crud/pedidos/pedidos_detalle.html', {'pedido': pedido})

def editar_pedido(request, id_pedido):
    pedido = get_object_or_404(Pedido, id_pedido=id_pedido)
    detalles = PedidoItem.objects.filter(pedido=pedido)
    
    if request.method == 'POST':
        estado_anterior = pedido.estado
        nuevo_estado = request.POST.get('estado')
        
        if nuevo_estado and nuevo_estado != estado_anterior:
            pedido.estado = nuevo_estado
            pedido.save()
            # Notificacion VERDE para editar
            messages.success(
                request, 
                f'Pedido #{pedido.id_pedido}: Estado cambiado de "{estado_anterior}" a "{nuevo_estado}"'
            )
        elif nuevo_estado == estado_anterior:
            messages.warning(
                request, 
                f'No se detectaron cambios en el estado del pedido #{pedido.id_pedido}'
            )
        
        return redirect('pedidos:lista')
    
    context = {
        'pedido': pedido,
        'detalles': detalles,
    }
    return render(request, 'roles/admin/Crud/pedidos/pedidos_editar.html', context)


def eliminar_pedido(request, id_pedido):
    pedido = get_object_or_404(Pedido, id_pedido=id_pedido)
    if request.method == 'POST':
        pedido_id_temp = pedido.id_pedido
        pedido.delete()
        # Notificacion ROJA para eliminar
        messages.error(
            request, 
            f'Pedido #{pedido_id_temp} eliminado exitosamente'
        )
        return redirect('pedidos:lista')
    return render(request, 'roles/admin/Crud/pedidos/pedidos_eliminar.html', {'pedido': pedido})


def eliminar_pedido(request, id_pedido):
    pedido = get_object_or_404(Pedido, id_pedido=id_pedido)

    if request.method == 'POST':
        pedido_id_temp = pedido.id_pedido

        pedido.delete()

        messages.success(
            request,
            f"Pedido #{pedido_id_temp} eliminado exitosamente",
            extra_tags='delete'
        )

        return redirect('pedidos:lista')

    return render(
        request,
        'roles/admin/Crud/pedidos/pedidos_eliminar.html',
        {'pedido': pedido}
    )
# ==================== ESTADÍSTICAS ====================

def estadisticas_pedidos(request):
    pedidos = Pedido.objects.all()
    total_ventas = pedidos.aggregate(Sum('total'))['total__sum'] or 0
    promedio_pedido = pedidos.aggregate(Avg('total'))['total__avg'] or 0
    conteo_pedidos = pedidos.count()

    context = {
        'total_ventas': total_ventas,
        'promedio_pedido': promedio_pedido,
        'conteo_pedidos': conteo_pedidos,
        'fecha': datetime.now()
    }
    
    return render(request, 'roles/admin/Crud/pedidos/pedidos_estadisticas.html', context)

# ==================== EXPORTACIONES ====================

def exportar_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedidos_fragata.pdf"'
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    p.drawString(100, 750, "Reporte de Pedidos - La Fragata Giratoria")
    p.save()
    response.write(buffer.getvalue())
    buffer.close()
    return response

def exportar_excel(request):
    return HttpResponse("Lógica de Excel para pedidos pendiente")


def generar_excel_seleccionados(ids):
    """Genera un archivo Excel con los pedidos seleccionados."""
    # Obtener pedidos
    pedidos = Pedido.objects.filter(id_pedido__in=ids).order_by('-fecha')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos Seleccionados"
    
    # Estilos
    fondo_negro = PatternFill(start_color="0F0F0F", end_color="0F0F0F", fill_type="solid")
    dorado = "D4AF37"
    borde_dorado = Border(
        left=Side(style="thin", color=dorado),
        right=Side(style="thin", color=dorado),
        top=Side(style="thin", color=dorado),
        bottom=Side(style="thin", color=dorado),
    )
    
    # Encabezados
    headers = ["ID Pedido", "Fecha", "Cliente", "Usuario", "Método Pago", "Estado", "Total"]
    ws.append(headers)
    
    # Estilo de encabezados
    for col in ws[1]:
        col.fill = fondo_negro
        col.font = Font(color=dorado, bold=True)
        col.alignment = Alignment(horizontal="center")
        col.border = borde_dorado
    
    # Datos
    for pedido in pedidos:
        cliente_nombre = pedido.id_cliente.nombre if pedido.id_cliente else "General"
        usuario_nombre = pedido.id_usuario.nombre_usuario if pedido.id_usuario else "N/A"
        metodo_nombre = pedido.id_metodo_pago.nombre_metodo if pedido.id_metodo_pago else "N/A"
        
        ws.append([
            pedido.id_pedido,
            pedido.fecha.strftime("%d/%m/%Y") if pedido.fecha else "",
            cliente_nombre,
            usuario_nombre,
            metodo_nombre,
            pedido.estado,
            float(pedido.total)
        ])
    
    # Estilo de datos
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fondo_negro
            cell.font = Font(color="FFFFFF")
            cell.border = borde_dorado
            cell.alignment = Alignment(horizontal="center")
    
    # Ancho de columnas
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="pedidos_seleccionados.xlsx"'
    wb.save(response)
    return response


class ListaPedidos(ListView):
    model = Pedido
    template_name = 'roles/admin/Crud/pedidos/pedidos.html'
    context_object_name = 'pedidos'
    paginate_by = 10
    
def procesar_seleccionados(request):
    if request.method == "POST":
        ids = request.POST.getlist('pedido_ids') # Captura los valores de los checkboxes
        action = request.POST.get('action')      # Identifica si es 'delete' o 'export'

        if not ids:
            return redirect('pedidos:lista')

        if action == 'delete':
            # Filtrar y eliminar
            Pedido.objects.filter(id_pedido__in=ids).delete()
            messages.success(request, "Selección eliminada correctamente.")
            
        elif action == 'export':
            # Aquí llamas a tu lógica de exportación pasando los IDs
            return generar_excel_seleccionados(ids)

    return redirect('pedidos:lista')